from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font
import tempfile
import base64
from PIL import Image
from io import BytesIO
import logging
import os
from django.conf import settings
from .views import *

logger = logging.getLogger(__name__)

def generate_timesheet_pdf(user, location_name, shifts, last_monday, last_sunday):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=timesheet_{user.first_name}_{user.last_name}_{location_name}_{last_monday.strftime("%Y-%m-%d")}_to_{last_sunday.strftime("%Y-%m-%d")}.pdf'
    
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Offset everything lower on the page
    y_offset = 100
    
    # Draw the logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'bblogo.png')
    try:
        logo = Image.open(logo_path)
        logo_width, logo_height = logo.size
        aspect_ratio = logo_width / logo_height
        logo_width = (width - 40) * 0.6  # Scale down the logo width by 20%
        logo_height = logo_width / aspect_ratio
        p.drawImage(logo_path, 20, height - logo_height - 20, width=logo_width, height=logo_height)
    except FileNotFoundError:
        logger.error(f"Logo file not found at {logo_path}")
        return HttpResponse("Logo file not found", status=404)
    
    # Draw the header
    p.setFont("Helvetica-Bold", 24)
    p.setFillColorRGB(0, 0, 0.5)  # Dark blue color
    p.drawString(50, height - logo_height - 50 - y_offset, "Temporary Agency Worker – Time Sheet")
    p.setFont("Helvetica", 12)
    p.setFillColorRGB(0, 0, 0)  # Black color
    p.drawString(50, height - logo_height - 80 - y_offset, f"Agency: {user.first_name} {user.last_name}")
    p.drawString(50, height - logo_height - 100 - y_offset, f"Location: {location_name}")
    p.drawString(50, height - logo_height - 120 - y_offset, f"Week: {last_monday.strftime('%d/%m/%Y')} to {last_sunday.strftime('%d/%m/%Y')}")
    
    # Add more spacing between the header and the column headers
    y = height - logo_height - 170 - y_offset
    headers = ["Date", "Start Time", "End Time", "Hours Done", "Sleep In", "Signature", "Signed By"]
    x_positions = [50, 120, 190, 260, 330, 400, 470]
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)
    
    # Add more spacing between the column headers and the first entry
    y -= 40
    total_hours = 0
    total_sleep_in = 0
    row_height = 40  # Increase the row height for better spacing
    for shift in shifts:
        hours_done = (datetime.combine(date.min, shift.end_time) - datetime.combine(date.min, shift.start_time)).seconds / 3600
        total_hours += hours_done
        total_sleep_in += 1 if shift.sleep_in else 0
        
        p.drawString(50, y, shift.date.strftime('%d/%m/%Y'))
        p.drawString(120, y, shift.start_time.strftime('%H:%M'))
        p.drawString(190, y, shift.end_time.strftime('%H:%M'))
        p.drawString(260, y, f"{hours_done:.2f}")
        p.drawString(330, y, "Yes" if shift.sleep_in else "No")
        
        if shift.signature:
            try:
                # Decode the base64 signature
                signature_data = shift.signature.split(',')[1]
                signature_image = Image.open(BytesIO(base64.b64decode(signature_data)))
                signature_image = signature_image.convert("RGBA")  # Ensure the image has an alpha channel
                
                # Create a white background image
                white_bg = Image.new("RGB", signature_image.size, "white")
                white_bg.paste(signature_image, (0, 0), signature_image)
                
                # Increase the resolution of the signature image
                signature_image.thumbnail((50, 20))  # Adjust the thumbnail size to fit within the row
                
                # Create a temporary file for the signature image
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_file:
                    white_bg.save(tmp_file.name)
                    p.drawImage(tmp_file.name, 400, y - 10, width=50, height=20)
            except Exception as e:
                logger.error(f"Error processing signature for shift {shift.id}: {e}")
        
        signed_by = shift.signed_by if shift.signed_by else "N/A"
        p.drawString(470, y, signed_by)
        
        y -= row_height  # Adjust the spacing for the next shift
    
    # Add more spacing between the last entry and the totals
    y -= 40
    p.drawString(50, y, "Total")
    p.drawString(260, y, f"{total_hours:.2f}")
    p.drawString(330, y, f"{total_sleep_in}")
    
    # Add the agency name and signature
    y -= 60  # Add some spacing before the agency section
    p.drawString(50, y, "Agency Name:")
    p.drawString(150, y, f"{user.first_name} {user.last_name}")
    
    # Retrieve the user's profile
    profile = WorkerProfile.objects.get(user=user)
    
    if profile.signature:
        try:
            # Decode the base64 signature
            signature_data = profile.signature.split(',')[1]
            signature_image = Image.open(BytesIO(base64.b64decode(signature_data)))
            signature_image = signature_image.convert("RGBA")  # Ensure the image has an alpha channel
            
            # Create a white background image
            white_bg = Image.new("RGB", signature_image.size, "white")
            white_bg.paste(signature_image, (0, 0), signature_image)
            
            # Increase the resolution of the signature image
            signature_image.thumbnail((100, 40))  # Adjust the thumbnail size to fit within the row
            
            # Create a temporary file for the signature image
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_file:
                white_bg.save(tmp_file.name)
                p.drawImage(tmp_file.name, 50, y - 50, width=100, height=40)
        except Exception as e:
            logger.error(f"Error processing user signature for user {user.id}: {e}")
    
    p.showPage()
    p.save()
    
    return response


def generate_timesheet_excel(user, location_name, shifts, last_monday, last_sunday):
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Sheet"
    
    # Add the header
    ws.merge_cells('A1:G1')
    # ws['A1'] = "Temporary Agency Worker – Time Sheet"
    # ws['A1'].font = Font(size=24, bold=True, color="0000FF")
    
    ws['A2'] = f"Worker: {user.first_name} {user.last_name}"
    ws['A3'] = f"Week: {last_monday.strftime('%Y-%m-%d')} to {last_sunday.strftime('%Y-%m-%d')}"
    
    # Add the table headers
    headers = ["Date", "Start Time", "End Time", "Hours Done", "Sleep In", "Signature", "Signed By"]
    ws.append(headers)
    
    total_hours = 0
    total_sleep_in = 0
    for shift in shifts:
        hours_done = (datetime.combine(date.min, shift.end_time) - datetime.combine(date.min, shift.start_time)).seconds / 3600
        total_hours += hours_done
        total_sleep_in += 1 if shift.sleep_in else 0
        
        row = [
            shift.date.strftime('%Y-%m-%d'),
            shift.start_time.strftime('%H:%M'),
            shift.end_time.strftime('%H:%M'),
            f"{hours_done:.2f}",
            "Yes" if shift.sleep_in else "No",
            "Signature",  # Placeholder for signature
            shift.signed_by if shift.signed_by else "N/A"
        ]
        ws.append(row)
    
    # Add the totals row
    ws.append(["Total", "", "", f"{total_hours:.2f}", f"{total_sleep_in}", "", ""])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=timesheet_{user.first_name}_{user.last_name}_{location_name}_{last_monday.strftime("%Y-%m-%d")}_to_{last_sunday.strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)
    
    return response